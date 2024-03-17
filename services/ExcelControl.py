import os
import openpyxl # Module  for the excel operations
from services.JsonControl import JsonControl
import sys

class ExcelControl:
    """
    This class is used for the export to excel process.
    It exports the scene names to the excel file

    """
    
    def __init__(self):

        self.jsonObj = JsonControl()
        self.path = self.jsonObj.checkExcelPath()

        if self.jsonObj.check_rowCounter() == 2:
            
            # Workbook() takes one, non-optional, argument
            # which is the filename that we want to create.
            self.workbook = openpyxl.Workbook()
            # The workbook object is then used to add new
            # worksheet via the add_worksheet() method.
            self.worksheet = self.workbook.active # Select worksheet as default
            self.worksheet.title = "Objects-Distances" # Set worksheet title

            # It creates Scene Table with Scene Name and Tag Columns
            self.create_table()
        else:
            self.workbook = openpyxl.load_workbook(os.path.join(
                os.getcwd(),
                "services",
                "object-distance.xlsx"
            ))



        self.workbook.close()

# Yeniden açarak yazma modunda açın
        self.workbook = openpyxl.load_workbook(os.path.join(
                os.getcwd(),
                "services",
                "object-distance.xlsx"
            ))
        self.worksheet = self.workbook.active



    
    def create_table(self):
        
        # Use the worksheet object to set values
        self.worksheet["A1"].value = 'ObjectName'
        self.worksheet["B1"].value = 'Size'
        self.worksheet["C1"].value = 'Region'
        self.worksheet["D1"].value = 'isClose'
        self.workbook.save(os.path.join(
            os.getcwd(),
            "services",
            "object-distance.xlsx"
        )) # Finally, save the excel file

    
    def add_info_to_table(self,row,cell,info):
        # Args: filename-> Filename to add excel table
        self.worksheet[f"{cell}{row}"].value = info # Set cell value
        self.workbook.save(os.path.join(
            os.getcwd(),
            "services",
            "object-distance.xlsx"
        )) # Finally, save the excel file
 
    
    def close_excel(self):
        self.workbook.close()
        

